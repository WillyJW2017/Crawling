import openpyxl
import datetime

file_path = 'C:\\Willy\\Report\\Shenzhen Report 01 July to 31 July.xlsx'
sheet_name = 'Sheet0'

workbook = openpyxl.load_workbook(file_path)
sheet = workbook[sheet_name]

min_num = 2
max_num = 2

for i in range(2, sheet.max_row+1):
    last_name1 = sheet.cell(row=i, column=1).value
    first_name1 = sheet.cell(row=i, column=2).value
    date_value1 = sheet.cell(row=i, column=3).value
    time_value1 = sheet.cell(row=i, column=4).value

    if time_value1 is None:
        print(i)
        exit()
    else:
        time_value1 = datetime.datetime.strptime(time_value1, '%H:%M:%S')

    last_name2 = sheet.cell(row=i+1, column=1).value
    first_name2 = sheet.cell(row=i+1, column=2).value
    date_value2 = sheet.cell(row=i+1, column=3).value
    time_value2 = sheet.cell(row=i+1, column=4).value


    if time_value2 is None:
        exit()
    else:
        time_value2 = datetime.datetime.strptime(time_value2, '%H:%M:%S')

    col_value_thr1 = last_name1+first_name1+date_value1
    col_value_two1 = last_name1+first_name1

    col_value_thr2 = last_name2 + first_name2 + date_value2
    col_value_two2 = last_name2 + first_name2

    if col_value_thr1 == col_value_thr2:
        max_num = i+1
        if time_value2 < time_value1:
            min_num = i+1

        if sheet.cell(row=min_num, column=5).value is None:
            sheet.cell(row=min_num, column=5, value='Y')


    elif col_value_two1 != col_value_two2:
        min_num = i+1


    sheet.cell(row=max_num, column=5, value='Y')
    sheet.cell(row=min_num, column=5, value='Y')

    if (col_value_thr2 is None) or (col_value_thr2 == ''):
        sheet.cell(row=i, column=5, value='Y')

    workbook.save(filename=file_path)